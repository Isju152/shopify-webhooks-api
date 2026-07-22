from flask import Flask, request, jsonify, render_template_string, send_file
from datetime import datetime, timedelta, timezone
import os
from dotenv import load_dotenv
import requests
import logging
from openpyxl import Workbook
import cloudscraper
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


load_dotenv()

app = Flask(__name__)

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Variables
DROPI_TOKEN = os.getenv('DROPI_TOKEN', "") # Nuevo Token de Dropi


# ==========================================
# NUEVO MOTOR DE ESTADOS CON DROPI
# ==========================================

def obtener_ordenes_dropi():
    """Se conecta a Dropi evadiendo firewalls con Cloudscraper"""
    
    if not DROPI_TOKEN:
        logger.error("❌ No se encontró la variable DROPI_TOKEN en el entorno.")
        return []

    # Limpiamos el token por si acaso
    token_limpio = DROPI_TOKEN.strip().strip('"').strip("'")

    # Fechas (últimos 30 días)
    hoy = datetime.now()
    hace_30_dias = hoy - timedelta(days=30)
    str_hoy = hoy.strftime('%Y-%m-%d')
    str_pasado = hace_30_dias.strftime('%Y-%m-%d')

    url = "https://api.dropi.mx/api/orders/myorders"
    
    params = {
        "orderBy": "id",
        "orderDirection": "desc",
        "result_number": 1000,
        "start": 0,
        "user_id": 136493,
        "from": str_pasado,
        "until": str_hoy
    }
    
    # Dejamos que Cloudscraper maneje el User-Agent, solo pasamos Origin, Referer y Auth
    headers = {
        "Authorization": f"Bearer {token_limpio}",
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://app.dropi.mx",
        "Referer": "https://app.dropi.mx/"
    }
    
    try:
        logger.info("📡 Conectando a Dropi usando Cloudscraper...")
        
        # 1. Creamos la instancia que simula ser Chrome en Windows
        scraper = cloudscraper.create_scraper(
            browser={
                'browser': 'chrome',
                'platform': 'windows',
                'desktop': True
            }
        )
        
        # 2. Hacemos la petición con el scraper en vez de requests
        response = scraper.get(url, headers=headers, params=params)
        
        logger.info(f"📊 STATUS DROPI: {response.status_code}")
        logger.info(f"📝 TEXTO DROPI: {response.text[:300]}")
        
        if response.status_code != 200:
            logger.error(f"❌ Error HTTP {response.status_code} al conectar con Dropi.")
            return []
            
        data_json = response.json()
        
        if isinstance(data_json, list):
            return data_json
        elif isinstance(data_json, dict):
            if 'objects' in data_json:
                return data_json['objects']
            elif 'data' in data_json:
                return data_json['data']
        
        return []
            
    except Exception as e:
        logger.error(f"❌ Error conectando con Dropi mediante cloudscraper: {e}")
        return []

def analizar_estancamiento_dropi(ordenes_dropi):
    """Filtra las órdenes de Dropi que llevan más de 24 horas sin moverse"""
    ordenes_estancadas = []
    
    # Estados finales en los que NO nos importa si llevan más de 24 hrs
    estados_completados = [
        'DEVOLUCION EN PROCESO',
        'DEVOLUCION', 
        'ENTREGADO', 
        'CANCELADO', 
        'PAQUETE EN DEVOLUCION'
    ]
    
    now = datetime.now(timezone.utc)

    for orden in ordenes_dropi:
        # Extraemos los datos según la estructura de Dropi
        estado_actual = str(orden.get('status_name', '')).lower()
        
        # Si la orden ya terminó su ciclo, la ignoramos
        if estado_actual in estados_completados:
            continue
            
        # Obtenemos la fecha de última actualización
        updated_at = orden.get('updated_at')
        if not updated_at:
            continue
            
        try:
            # Parseamos la fecha de Dropi
            updated_datetime = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
            horas_sin_actualizar = (now - updated_datetime).total_seconds() / 3600
            
            if horas_sin_actualizar >= 24:
                orden['horas_estancada'] = round(horas_sin_actualizar, 1)
                ordenes_estancadas.append(orden)
        except Exception as e:
            logger.warning(f"Error procesando fecha para orden {orden.get('id')}: {e}")
            
    return ordenes_estancadas

def generar_reporte_excel_dropi(ordenes_estancadas):
    """Genera el Excel basándose PURAMENTE en la data de Dropi"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Alertas Dropi > 24h"
    
    header_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers_excel = [
        "Guía / Tracking",
        "Cliente",
        "Teléfono",
        "Estado en Dropi",
        "Última Actualización",
        "Horas Estancada",
        "Transportadora",
        "Ciudad"
    ]
    ws.append(headers_excel)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        
    for orden in ordenes_estancadas:
        row = [
            orden.get('tracking', 'Sin guía'),
            orden.get('client_name', 'Desconocido'),
            orden.get('client_phone', 'Sin teléfono'),
            orden.get('status_name', 'Sin estado').upper(),
            orden.get('updated_at', ''),
            f"{orden.get('horas_estancada')} hrs",
            orden.get('carrier', 'No asignada'),
            orden.get('city', 'No especificada')
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 20

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

# ==========================================
# ENDPOINTS DEL DASHBOARD
# ==========================================

@app.route('/check-stale-orders', methods=['GET'])
def check_stale_orders():
    """Ahora chequea directamente en Dropi"""
    logger.info("🔍 Chequeando Dropi por órdenes estancadas...")
    
    ordenes_dropi = obtener_ordenes_dropi()
    
    if not ordenes_dropi:
        return jsonify({'error': 'No se pudo conectar a Dropi o no hay órdenes'}), 500
        
    estancadas = analizar_estancamiento_dropi(ordenes_dropi)
    
    return jsonify({
        'status': 'ok',
        'stale_orders': len(estancadas)
    }), 200

@app.route('/descargar-reporte-estancadas', methods=['GET'])
def descargar_reporte_estancadas():
    """Descarga el Excel basado en Dropi"""
    ordenes_dropi = obtener_ordenes_dropi()
    estancadas = analizar_estancamiento_dropi(ordenes_dropi)
    
    excel_file = generar_reporte_excel_dropi(estancadas)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Dropi_Estancadas_{datetime.now().strftime("%Y-%m-%d_%H-%M")}.xlsx'
    )

# --- RUTAS DE DASHBOARD Y WEBHOOKS SHOPIFY ORIGINALES ---
@app.route('/dashboard', methods=['GET'])
def dashboard():
    # Tu mismo HTML del dashboard, sin modificaciones mayores.
    # Solo cambié el título para que diga "Chequear Órdenes en Dropi"
    html = """
    <!DOCTYPE html>
    <html lang="es">
    <!-- Mismo CSS que ya tenías -->
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Dashboard - Amarela Webhooks</title>
        <style>
            body { font-family: 'Segoe UI', sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; display: flex; justify-content: center; align-items: center; padding: 20px; }
            .container { background: white; border-radius: 10px; box-shadow: 0 10px 40px rgba(0, 0, 0, 0.2); padding: 40px; max-width: 500px; width: 100%; }
            h1 { color: #333; margin-bottom: 10px; text-align: center; }
            .subtitle { color: #666; text-align: center; margin-bottom: 40px; font-size: 14px; }
            .buttons-group { display: flex; flex-direction: column; gap: 15px; }
            button { width: 100%; padding: 15px; border: none; border-radius: 5px; font-size: 16px; font-weight: bold; cursor: pointer; transition: transform 0.2s; }
            .btn-check { background: linear-gradient(135deg, #FF9900 0%, #FF6600 100%); color: white; }
            .btn-report { background: linear-gradient(135deg, #27ae60 0%, #229954 100%); color: white; }
            .result { margin-top: 30px; padding: 20px; background: #f8f9fa; border-radius: 5px; display: none; }
            .result.show { display: block; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📊 Dashboard Logístico (Dropi)</h1>
            <p class="subtitle">Monitoreo de órdenes estancadas en transportadora</p>
            <div class="buttons-group">
                <button class="btn-check" id="checkButton" onclick="checkStaleOrders()">
                    🔍 Chequear Dropi (>24h sin movimiento)
                </button>
                <button class="btn-report" onclick="descargarReporte()">
                    📊 Descargar Reporte Excel
                </button>
            </div>
            <div class="result" id="result">
                <h3>Resultado:</h3>
                <div id="resultContent"></div>
            </div>
        </div>
        <script>
            async function checkStaleOrders() {
                const button = document.getElementById('checkButton');
                const result = document.getElementById('result');
                const resultContent = document.getElementById('resultContent');
                
                button.disabled = true;
                button.innerHTML = 'Ejecutando chequeo en Dropi...';
                
                try {
                    const response = await fetch('/check-stale-orders');
                    const data = await response.json();
                    
                    if (!response.ok) {
                        resultContent.innerHTML = `<p style="color: red;">❌ Error: ${data.error}</p>`;
                        result.classList.add('show');
                        return;
                    }
                    
                    if (data.stale_orders === 0) {
                        resultContent.innerHTML = `<p style="color: green;">✅ Todo fluyendo. No hay órdenes estancadas.</p>`;
                    } else {
                        resultContent.innerHTML = `<p style="color: red;">⚠️ <strong>${data.stale_orders}</strong> órdenes detectadas sin movimiento por más de 24 horas.</p>`;
                    }
                    result.classList.add('show');
                } catch (error) {
                    resultContent.innerHTML = `<p style="color: red;">❌ Error de conexión</p>`;
                    result.classList.add('show');
                } finally {
                    button.disabled = false;
                    button.innerHTML = '🔍 Chequear Dropi (>24h sin movimiento)';
                }
            }
            function descargarReporte() {
                window.location.href = '/descargar-reporte-estancadas';
            }
        </script>
    </body>
    </html>
    """
    return render_template_string(html)

@app.route('/', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'message': 'Servidor Amarela activo'})

# (Aquí puedes dejar tus funciones originales de @app.route('/webhooks/orders/create') 
# si aún quieres que te lleguen correos cuando alguien compra en la tienda)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)